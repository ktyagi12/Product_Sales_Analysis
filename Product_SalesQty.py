'''
PROBLEM STATEMENT:
1.	Create an excel file with headers product and Qty.
  	Ask user for num of products
	Input from user: product name and qty.
	Save the file

2.	Read the same file
	Create a dictionary of product data
	Print the data in tabular format

3.	Plot the various graphs
'''

import openpyxl as op
import pandas as pd
import matplotlib.pyplot as plt

path = r"D:\KTyagi\KT\Python_Workspace\DataScience\Assignments\Product_SalesQty.xlsx"

workbook = op.load_workbook(path)
sheet_obj = workbook.active
data = {}

def create_header(path,wk_book,sheet_obj):
	cell_obj= sheet_obj.cell(row = 1, column = 1)
	cell_obj.value = 'PRODUCT'

	cell_obj1 = sheet_obj.cell(row = 1, column = 2)
	cell_obj1.value = 'QTY'

	workbook.save(path)

def insert_data():
	num_products = int(raw_input('How many products: '))
	for i in range(num_products):
		cell_obj = sheet_obj.cell(row = (i+2), column = 1)
		cell_obj1 = sheet_obj.cell(row = (i+2), column = 2)
		prod_name = raw_input('Product Name: ')
		sales_qty = int(raw_input('Sales Qty: '))
		cell_obj.value = prod_name
		cell_obj1.value = sales_qty

		workbook.save(path)

def create_dict():
	for i in range(2,sheet_obj.max_row + 1):
		cell_obj = sheet_obj.cell(row = i, column = 1)
		if cell_obj.value not in data:
			data[cell_obj.value] = sheet_obj.cell(row = i,column =2).value
		else:
			pass
	print 'Data: ', data



def create_series(data):
	series = pd.Series(data)
	print 'Series: \n',series
	series.plot(kind = 'bar',color = 'green')
	plt.show()

	series.plot(kind = 'hist',color = 'red')
	plt.show()

	series.plot(kind = 'pie')
	plt.show()

	series.plot(kind = 'barh',color = 'blue')
	plt.show()

	series.plot(kind = 'line',color = 'purple')
	plt.show()
	
	'''series.plot(kind = 'kde')
	   plt.show() 			 * It requires scipy library'''

create_header(path,workbook,sheet_obj)
insert_data()
create_dict()
create_series(data)